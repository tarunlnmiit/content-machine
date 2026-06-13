#!/usr/bin/env python3
"""
Generate an annual social media posting tracker spreadsheet.

Usage:
  python3 scripts/generate_posting_tracker.py              # current year
  python3 scripts/generate_posting_tracker.py --year 2026
"""

import argparse
import datetime
import json
import os
import sys
from collections import defaultdict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Colors ────────────────────────────────────────────────────────────────────
NICHE_COLORS = {"DS": "DDEEFF", "Life": "DDFFDD", "Poetry": "EED9FF"}
HEADER_BG    = "2D3748"
SUBHDR_BG    = "E2E8F0"

DAYS     = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_NUM  = {d: i + 1 for i, d in enumerate(DAYS)}  # 1=Mon … 7=Sun (ISO)
MONTHS   = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

NICHE_MARKERS = {
    "_data_science_tech_": "DS",
    "_life_self_dev_":     "Life",
    "_poetry_quotes_":     "Poetry",
}


# ── Schedule ──────────────────────────────────────────────────────────────────

LONGFORM_PLATFORMS = [
    ("DS",     "Medium",  "Article"),
    ("DS",     "YouTube", "Video"),
    ("Life",   "Medium",  "Article"),
    ("Life",   "YouTube", "Video"),
    ("Poetry", "Medium",  "Article"),
    ("Poetry", "YouTube", "Video"),
]

# (day, time, niche, platform, format)
WEEKLY = [
    # Monday
    ("Mon", "08:00 AM", "DS",     "LinkedIn",     "Post"),
    ("Mon", "12:00 PM", "DS",     "Twitter",      "Post"),
    ("Mon", "04:00 PM", "DS",     "FB/IG",        "Carousel"),
    ("Mon", "06:00 PM", "DS",     "LinkedIn",     "Poll"),
    ("Mon", "07:00 PM", "Life",   "Meta Threads", "Native Text Post"),
    # Tuesday
    ("Tue", "08:00 AM", "Life",   "LinkedIn",     "Post"),
    ("Tue", "12:00 PM", "DS",     "Twitter",      "Poll (manual)"),
    ("Tue", "06:00 PM", "Life",   "LinkedIn",     "Poll"),
    ("Tue", "08:30 PM", "Poetry", "FB/IG",        "Post"),
    ("Tue", "09:00 PM", "DS",     "Twitter",      "Twitter Thread"),
    # Wednesday
    ("Wed", "06:00 AM", "DS",     "LinkedIn",     "Slide Deck"),
    ("Wed", "06:00 AM", "Poetry", "FB/IG",        "Carousel"),
    ("Wed", "08:00 AM", "Poetry", "LinkedIn",     "Post"),
    ("Wed", "12:00 PM", "Life",   "Twitter",      "Post"),
    ("Wed", "06:00 PM", "Poetry", "LinkedIn",     "Poll"),
    ("Wed", "08:00 PM", "DS",     "Meta Threads", "Native Text Post"),
    # Thursday
    ("Thu", "08:00 AM", "Life",   "LinkedIn",     "Slide Deck"),
    ("Thu", "10:00 AM", "Life",   "FB/IG",        "Carousel"),
    ("Thu", "12:00 PM", "Life",   "Twitter",      "Poll (manual)"),
    ("Thu", "07:00 PM", "Poetry", "Meta Threads", "Native Text Post"),
    ("Thu", "09:00 PM", "Life",   "Twitter",      "Twitter Thread"),
    # Friday
    ("Fri", "08:00 AM", "Poetry", "LinkedIn",     "Slide Deck"),
    ("Fri", "10:00 AM", "Life",   "FB/IG",        "Post"),
    ("Fri", "12:00 PM", "Poetry", "Twitter",      "Post"),
    # Saturday
    ("Sat", "02:00 PM", "DS",     "FB/IG",        "Post"),
    ("Sat", "09:00 PM", "Poetry", "Twitter",      "Twitter Thread"),
    ("Sat", "12:00 PM", "Poetry", "Twitter",      "Poll (manual)"),
]

# (platform, time, niche, format) — repeated every day Mon–Sun
# DS Clip Short and Remotion Reels: 2× per day = 14 per week
DAILY = [
    # FB/IG
    ("FB/IG",        "07:00 AM", "Life",   "Clip Short Reel"),
    ("FB/IG",        "09:00 AM", "DS",     "Remotion Reel"),
    ("FB/IG",        "10:00 AM", "Poetry", "Remotion Reel"),
    ("FB/IG",        "01:00 PM", "DS",     "Clip Short Reel"),
    ("FB/IG",        "03:00 PM", "DS",     "Remotion Reel"),
    ("FB/IG",        "09:00 PM", "DS",     "Clip Short Reel"),
    ("FB/IG",        "09:00 PM", "Life",   "Remotion Reel"),
    ("FB/IG",        "10:00 PM", "Poetry", "Clip Short Reel"),
    # LinkedIn
    ("LinkedIn",     "07:00 AM", "Life",   "Clip Short Reel"),
    ("LinkedIn",     "09:00 AM", "Poetry", "Clip Short Reel"),
    ("LinkedIn",     "10:00 AM", "DS",     "Clip Short Reel"),
    ("LinkedIn",     "02:00 PM", "DS",     "Remotion Reel"),
    ("LinkedIn",     "04:00 PM", "DS",     "Clip Short Reel"),
    ("LinkedIn",     "06:00 PM", "Life",   "Remotion Reel"),
    ("LinkedIn",     "07:00 PM", "Poetry", "Remotion Reel"),
    ("LinkedIn",     "08:00 PM", "DS",     "Remotion Reel"),
    # Twitter
    ("Twitter",      "11:00 AM", "DS",     "Clip Short Reel"),
    ("Twitter",      "01:00 PM", "Life",   "Clip Short Reel"),
    ("Twitter",      "02:00 PM", "Poetry", "Clip Short Reel"),
    ("Twitter",      "03:00 PM", "DS",     "Remotion Reel"),
    ("Twitter",      "05:00 PM", "DS",     "Clip Short Reel"),
    ("Twitter",      "07:00 PM", "DS",     "Remotion Reel"),
    ("Twitter",      "08:00 PM", "Life",   "Remotion Reel"),
    ("Twitter",      "09:00 PM", "Poetry", "Remotion Reel"),
    # YouTube Shorts — DS 2×/day, Life 1×/day, Poetry 1×/day
    ("YouTube Shorts", "08:00 AM", "DS",     "Remotion Reel"),
    ("YouTube Shorts", "10:00 AM", "Life",   "Clip Short Reel"),
    ("YouTube Shorts", "10:00 AM", "Poetry", "Remotion Reel"),
    ("YouTube Shorts", "11:00 AM", "DS",     "Clip Short Reel"),
    ("YouTube Shorts", "02:00 PM", "DS",     "Remotion Reel"),
    ("YouTube Shorts", "07:00 PM", "Life",   "Remotion Reel"),
    ("YouTube Shorts", "08:00 PM", "DS",     "Clip Short Reel"),
    ("YouTube Shorts", "09:00 PM", "Poetry", "Clip Short Reel"),
]


# ── Content data loading ──────────────────────────────────────────────────────

def detect_niche(slug: str) -> str | None:
    for marker, niche in NICHE_MARKERS.items():
        if marker in slug:
            return niche
    return None

def slug_to_title(slug: str) -> str:
    for marker in NICHE_MARKERS:
        if marker in slug:
            idx = slug.index(marker) + len(marker)
            return slug[idx:].replace("-", " ").title()
    return slug.replace("-", " ").title()

def _parse_date(ts: str | None) -> datetime.date | None:
    if not ts:
        return None
    return datetime.datetime.fromisoformat(ts).date()

def load_content_data(year: int, repo_root: str) -> dict:
    """Return {(wk, niche): {"title", "slug", "publish_dates": {platform: date}}}."""
    data = {}
    deriv_dir = os.path.join(repo_root, "content", "derivatives")
    if not os.path.isdir(deriv_dir):
        return data
    for week_dir in os.listdir(deriv_dir):
        if not week_dir.startswith(f"{year}-W"):
            continue
        wk = int(week_dir.split("-W")[1])
        week_path = os.path.join(deriv_dir, week_dir)
        for slug in os.listdir(week_path):
            niche = detect_niche(slug)
            if not niche:
                continue

            # Title from youtube_metadata.json, fallback to slug
            title = ""
            meta_path = os.path.join(week_path, slug, "youtube_metadata.json")
            if os.path.exists(meta_path):
                with open(meta_path, encoding="utf-8") as f:
                    title = json.load(f).get("title", "").strip()
            if not title:
                title = slug_to_title(slug)

            # Platform publish dates from schedule.json
            publish_dates: dict[str, datetime.date] = {}
            sched_path = os.path.join(week_path, slug, "schedule.json")
            if os.path.exists(sched_path):
                with open(sched_path, encoding="utf-8") as f:
                    sched = json.load(f)

                lf = sched.get("long_form", {})
                if lf.get("publish_at"):
                    publish_dates["YouTube"] = _parse_date(lf["publish_at"])

                blog = sched.get("blog", {})
                if blog.get("medium_publish_at"):
                    publish_dates["Medium"] = _parse_date(blog["medium_publish_at"])

                social = sched.get("social", {})
                _social_map = {
                    "linkedin_publish_at":  "LinkedIn",
                    "twitter_publish_at":   "Twitter",
                    "ig_fb_publish_at":     "FB/IG",
                    "threads_publish_at":   "Meta Threads",
                }
                for key, plat in _social_map.items():
                    if social.get(key):
                        publish_dates[plat] = _parse_date(social[key])

            data[(wk, niche)] = {
                "title":         title,
                "slug":          slug,
                "publish_dates": publish_dates,
            }
    return data


# ── Status preservation ───────────────────────────────────────────────────────

def load_existing_statuses(out_path: str) -> dict:
    """Read existing xlsx and return {(iso_week, day, date_str, time_str, niche, platform, fmt): status}."""
    statuses = {}
    if not os.path.exists(out_path):
        return statuses
    try:
        wb = load_workbook(out_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            col_map = {h: i for i, h in enumerate(header_row) if h}
            for row in ws.iter_rows(min_row=2, values_only=True):
                def get(name: str):
                    idx = col_map.get(name)
                    return row[idx] if idx is not None and idx < len(row) else None

                status = get("Status")
                if not status:
                    continue
                key = (
                    get("ISO Week"), get("Day"), get("Date"),
                    get("Time"), get("Niche"), get("Platform"), get("Format"),
                )
                statuses[key] = status
        wb.close()
    except Exception:
        pass  # Never crash on corrupt/locked file
    return statuses


# ── Row generation ────────────────────────────────────────────────────────────

def time_minutes(t: str) -> int:
    dt = datetime.datetime.strptime(t, "%I:%M %p")
    return dt.hour * 60 + dt.minute

def max_iso_week(year: int) -> int:
    return datetime.date(year, 12, 28).isocalendar().week

def week_label(week: int) -> str:
    return f"W{week:02d}"

def make_rows(year: int, content_data: dict) -> list:
    """Return list of row dicts for the full year starting W21."""
    rows = []
    max_wk = max_iso_week(year)

    for wk in range(21, max_wk + 1):
        wl = week_label(wk)
        mon_date = datetime.date.fromisocalendar(year, wk, 1)

        empty_info = {"title": "", "slug": "", "publish_dates": {}}
        week_info = {n: content_data.get((wk, n), empty_info) for n in ("DS", "Life", "Poetry")}

        # Long-form (same week as content)
        for niche, platform, fmt in LONGFORM_PLATFORMS:
            date = mon_date
            if date.year != year:
                continue
            pub = week_info[niche]["publish_dates"]
            posting_date = pub.get(platform) or date
            rows.append({
                "date":         date,
                "posting_date": posting_date,
                "iso_week":     wl,
                "slug":         week_info[niche]["slug"],
                "day":          "Mon",
                "time":         "—",
                "niche":        niche,
                "platform":     platform,
                "format":       fmt,
                "title":        week_info[niche]["title"],
                "status":       "Idea",
            })

        # Weekly social posts (next week)
        for day, time_str, niche, platform, fmt in WEEKLY:
            date = datetime.date.fromisocalendar(year, wk, DAY_NUM[day])
            if date.year != year:
                continue
            pub = week_info[niche]["publish_dates"]
            posting_date = pub.get(platform) or (date + datetime.timedelta(days=7))
            rows.append({
                "date":         date,
                "posting_date": posting_date,
                "iso_week":     wl,
                "slug":         week_info[niche]["slug"],
                "day":          day,
                "time":         time_str,
                "niche":        niche,
                "platform":     platform,
                "format":       fmt,
                "title":        week_info[niche]["title"],
                "status":       "Idea",
            })

        # Daily reels (next week)
        for day in DAYS:
            date = datetime.date.fromisocalendar(year, wk, DAY_NUM[day])
            if date.year != year:
                continue
            for platform, time_str, niche, fmt in DAILY:
                posting_date = date + datetime.timedelta(days=7)
                rows.append({
                    "date":         date,
                    "posting_date": posting_date,
                    "iso_week":     wl,
                    "slug":         week_info[niche]["slug"],
                    "day":          day,
                    "time":         time_str,
                    "niche":        niche,
                    "platform":     platform,
                    "format":       fmt,
                    "title":        week_info[niche]["title"],
                    "status":       "Idea",
                })

    rows.sort(key=lambda r: (r["posting_date"], r["date"], r["time"] == "—", r["time"]))
    return rows


# ── Spreadsheet writer ────────────────────────────────────────────────────────

def niche_fill(niche: str) -> PatternFill:
    return PatternFill("solid", fgColor=NICHE_COLORS.get(niche, "FFFFFF"))

def solid_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def bold_white(size: int = 10) -> Font:
    return Font(bold=True, color="FFFFFF", size=size)

def bold_dark(size: int = 10) -> Font:
    return Font(bold=True, color="1A202C", size=size)

HEADERS = ["ISO Week", "Slug", "Day", "Date", "Posting Date", "Time",
           "Niche", "Platform", "Format", "Content Title", "Status", "✓"]
COL_WIDTHS = [9, 30, 6, 13, 13, 11, 8, 15, 22, 48, 14, 4]

STATUS_OPTIONS = "Idea,In Progress,Editing,Published"


def write_sheet(ws, month_rows: list, year: int, month_idx: int) -> None:
    ws.freeze_panes = "A2"

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font  = bold_white()
        c.fill  = solid_fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 18

    for r_idx, row in enumerate(month_rows, start=2):
        fill = niche_fill(row["niche"])
        vals = [
            row["iso_week"],
            row.get("slug", ""),
            row["day"],
            row["date"].strftime("%-d %b %Y"),
            row["posting_date"].strftime("%-d %b %Y"),
            row["time"],
            row["niche"],
            row["platform"],
            row["format"],
            row["title"],
            row["status"],
            "",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.fill = fill

    last_row = len(month_rows) + 1

    ws.auto_filter.ref = f"A1:L{last_row}"

    # Status dropdown — column K (11)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"K2:K{last_row}"


def write_annual_tracker(year: int, repo_root: str) -> str:
    content_data = load_content_data(year, repo_root)
    all_rows = make_rows(year, content_data)

    out_dir  = os.path.join(repo_root, "output", "trackers")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"annual-tracker-{year}.xlsx")

    # Preserve manually-entered Status values from existing file
    statuses = load_existing_statuses(out_path)
    preserved = 0
    for row in all_rows:
        key = (
            row["iso_week"], row["day"],
            row["date"].strftime("%-d %b %Y"),
            row["time"], row["niche"], row["platform"], row["format"],
        )
        if key in statuses:
            row["status"] = statuses[key]
            preserved += 1

    # Group by posting_date month (only current year)
    by_month: dict[int, list] = defaultdict(list)
    for row in all_rows:
        if row["posting_date"].year == year:
            by_month[row["posting_date"].month].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    for m in range(1, 13):
        month_rows = by_month.get(m, [])
        if not month_rows:
            continue
        ws = wb.create_sheet(title=MONTHS[m - 1])
        write_sheet(ws, month_rows, year, m)

    wb.save(out_path)
    print(f"Saved → {out_path}")
    print(f"  Rows: {len(all_rows)} across {len(by_month)} monthly sheets")
    print(f"  Content items loaded: {len(content_data)}")
    print(f"  Statuses preserved: {preserved}")
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate annual posting tracker")
    parser.add_argument("--year", type=int, default=datetime.date.today().year)
    args = parser.parse_args()

    repo_root = os.path.join(os.path.dirname(__file__), "..")
    write_annual_tracker(args.year, os.path.abspath(repo_root))


if __name__ == "__main__":
    main()
